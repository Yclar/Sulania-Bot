from nonebot import on_command
from nonebot.rule import to_me
from nonebot.typing import T_State
from nonebot.adapters import Bot, Event, Message
import openpyxl
import os
from pathlib import Path
from nonebot.params import State, CommandArg

login = on_command("login", rule=to_me(), priority=5)


@login.handle()
async def handle_first_receive(bot: Bot, event: Event, state: T_State = State(), args: Message = CommandArg()):
    plain_text = args.extract_plain_text()
    text = plain_text.split(" ")
    for x in text:
        if x.find('luogu=') != -1:
            if x[x.find('luogu=') + 6:].isdigit():
                state['luogu_id'] = x[x.find('luogu=')+6:]
            else:
                login.reject("你的luogu_id有错误哦~")


@login.handle()
async def reply_and_send(bot: Bot, event: Event, state: T_State = State()):
    str_path = Path(os.getcwd())
    sx = openpyxl.load_workbook(str_path / 'sulania_bot' / 'user.xlsx')
    user_file = sx['Sheet1']
    for x in range(2, user_file.max_row+1):
        if user_file.cell(row=x, column=1).value == str(event.get_user_id()):
            await login.reject("你已经登录过了~")
            await login.finish()
    for r in range(2, user_file.max_row+1):
        if user_file.cell(row=r, column=1).value is None:
            user_file.cell(row=r, column=1).value = str(event.get_user_id())
            if state['luogu_id'] is not None:
                user_file.cell(row=r, column=2).value = state['luogu_id']
            break
    sx.save(str_path / 'sulania_bot' / 'user.xlsx')
    await login.send("成功！")
    await login.finish()
